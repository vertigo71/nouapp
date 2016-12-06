import logging, inspect
from random import randint

from django.core.management.base import BaseCommand, CommandError
from django.db.models import Max,Min

from openpyxl import load_workbook

from nouapp.models import Nou, Person, Quote

# Excel row where are written all nname's
ROWNAME= 1
# Excel col where are written all the dates
COLDATE= 1

# open Excel file 
# returns active   worksheet
def getws(file ):
    logger = logging.getLogger(__name__)
    logger.info("Starting {}".format( inspect.stack()[0][3]) )
   
    # open Excel File
    try:
        logger.info("  Loading = {}".format( file ) )
        wb = load_workbook(file)
        ws = wb.active
        logger.info("  Active Sheet = %s", ws.title )
        print("Active NOU sheet = ", ws.title )
    except:
        print("Error opening Excel file: %s\n\n" % file )
        raise
        
    return ws

def excel2database( ws ):
    logger = logging.getLogger(__name__)
    logger.info("Starting {}".format( inspect.stack()[0][3]) )

    # process Users
    rows = ws.get_squared_range(COLDATE+1,ROWNAME,ws.max_column,ROWNAME)
    for row in rows:
        for numcol,cell in enumerate( row , start = COLDATE+1 ):
            if cell.value and cell.value != '':
                if not Person.objects.filter( name = cell.value ):
                    # insert new Person
                    logger.info("  Inserting Person: row = {} : col = {} : person = {}".format( ROWNAME, numcol, cell.value ))
                    # To create and save an object in a single step, use the create() method.
                    Person.objects.create( name = cell.value )
            else:
                logger.info("  Person Error: row = {} : col = {} : cell = {}".format( ROWNAME, numcol, cell ))
  
    # Quote
    minpk = Quote.objects.all().aggregate(Min('pk'))['pk__min'] # None if empty
    maxpk = Quote.objects.all().aggregate(Max('pk'))['pk__max'] # None if empty
    logger.info( "  minpk = {} : maxpk = {}".format( minpk , maxpk ) )
    if not minpk or not maxpk:
        logger.error( "  minpk = {} : maxpk = {}".format( minpk , maxpk ) )
        raise CommandError("No quotes available" )
    
    # process Data
    rows = ws.get_squared_range(COLDATE+1,ROWNAME+1,ws.max_column,ws.max_row)
    for numrow,row in enumerate( rows, start=ROWNAME+1 ):

        if numrow > 5:
            break
               
        # get date per row
        cell = ws.cell( row=numrow ,column=COLDATE)
        if cell.is_date and cell.value:
            rowdate = cell.value.date()
        else:
            logger.info("  Row Date Error: row = {} : col = {} : cell = {}".format( numrow, COLDATE, cell.value ) )
            continue

        print( "Processing row = {} : date = {}".format( numrow , rowdate ) )
        logger.info( "  Processing row = {} : date = {}".format( numrow , rowdate ) )
        
        for numcol, cell in enumerate(row, start = COLDATE+1 ):
            if cell.value and cell.value != '':
                # get Person
                name = ws.cell( row=ROWNAME ,column=numcol).value
                try:
                    person = Person.objects.get( name = name )
                except Person.DoesNotExist:
                    raise CommandError('Person <{}> does not exist'.format(person) )

                if not Nou.objects.filter( person = person , date = rowdate ):
                    # insert new Nou record
                    # get a random Quote
                    pk = randint( minpk, maxpk )
                    quote = Quote.objects.filter( pk__gte=pk )[0] # obtain 1 random quote
                    logger.info("  Inserting Nou: {} : person = {} : date = {} : quote = {}".format( cell.value, name, rowdate, pk ) )
                    # To create and save an object in a single step, use the create() method.
                    Nou.objects.create( person = person , quote=quote, num = cell.value, date = rowdate )
            else:
                logger.info("  Nou Error: row = {} : col = {} : cell = {}".format( numrow, numcol, cell ) )
                    


class Command(BaseCommand):
    help = 'Loads the excel file into the database (first populate quotes)'

    def add_arguments(self, parser):
        parser.add_argument('excelfile', metavar='Excel file', type=str )

    def handle(self, *args, **options):
        logger = logging.getLogger(__name__)
        logger.info("------------------------------------------------------------" )
        logger.info("Starting {}".format( inspect.stack()[0][3]) )

        # open NOU and get active sheet
        logger.info("Excel File = {}".format( options['excelfile'] ) )
        ws = getws( options['excelfile'] )
        
        # insert Person, Nou
        excel2database( ws )
                
        self.stdout.write(self.style.SUCCESS('Successfully populated the database' ))
